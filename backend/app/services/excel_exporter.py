"""
Excel exporter: write the cleaned data + analysis tables into a single
workbook with one sheet per logical section.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Light styling used in every header row.
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    """Write a header row + body rows with light styling and auto column widths."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    for row in rows:
        ws.append(row)

    # Auto-fit column widths to the longest cell in each column.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                value = row[col_idx - 1]
                max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def _df_to_rows(df: pd.DataFrame) -> tuple[list[str], list[list[Any]]]:
    """Convert a DataFrame to (headers, list-of-row-lists) safely for openpyxl."""
    headers = [str(c) for c in df.columns]
    rows: list[list[Any]] = []
    for record in df.itertuples(index=False, name=None):
        row: list[Any] = []
        for value in record:
            if value is None or (isinstance(value, float) and pd.isna(value)):
                row.append("")
            elif isinstance(value, pd.Timestamp):
                row.append(value.strftime("%Y-%m-%d"))
            else:
                row.append(value)
        rows.append(row)
    return headers, rows


def export_to_excel(
    cleaned_df: pd.DataFrame,
    monthly_trend: list[dict[str, Any]],
    category_breakdown: list[dict[str, Any]],
    top_products: list[dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """
    Write a workbook with these sheets:
        - Cleaned Data
        - Monthly Summary
        - Category Summary
        - Top Products
    Returns the absolute path of the saved file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # 1. Cleaned Data sheet (replace the default sheet).
    ws_clean = wb.active
    ws_clean.title = "Cleaned Data"
    headers, rows = _df_to_rows(cleaned_df)
    _write_table(ws_clean, headers, rows)

    # 2. Monthly Summary sheet.
    ws_month = wb.create_sheet("Monthly Summary")
    _write_table(
        ws_month,
        ["Month", "Sales", "Expenses", "Profit"],
        [[r["month"], r["sales"], r["expenses"], r["profit"]] for r in monthly_trend],
    )

    # 3. Category Summary sheet.
    ws_cat = wb.create_sheet("Category Summary")
    _write_table(
        ws_cat,
        ["Category", "Amount"],
        [[r["category"], r["amount"]] for r in category_breakdown],
    )

    # 4. Top Products sheet.
    ws_top = wb.create_sheet("Top Products")
    _write_table(
        ws_top,
        ["Product", "Sales"],
        [[r["product"], r["sales"]] for r in top_products],
    )

    wb.save(output_path)
    return output_path.resolve()
